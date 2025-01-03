import openpyxl

#定义excel批量处理函数
def process_excel(input_path, output_path):
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(input_path)

    # 遍历所有工作簿
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # 遍历所有单元格
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value:
                    # 处理 . 符号
                    value = cell.value
                    while '.' in value:
                        dot_index = value.find('.')
                        if dot_index != -1 and dot_index + 4 < len(value):
                            value = value[:dot_index + 1] + value[dot_index + 2:dot_index + 4] + value[dot_index + 5:]
                        else:
                            break
                    
                    # 处理 ( 符号
                    count = 1
                    new_value = ""
                    for char in value:
                        if char == '(':
                            new_value += f"{count}{char}"
                            count += 1
                        else:
                            new_value += char
                    
                    # 更新单元格的值
                    cell.value = new_value

    # 保存修改后的 Excel 文件
    workbook.save(output_path)

# 指定输入输出路径
input_file_path = 'your_input_file.xlsx'
output_file_path = 'your_output_file.xlsx'

#执行excel批量处理函数
process_excel(input_file_path, output_file_path)
